#!/usr/bin/env python3
"""Fetch data from a public Google Sheet (XLSX export) and write JSON data files."""

import argparse
import datetime
import io
import json
import os
import re
import sys

import openpyxl
import requests


def cell_to_str(value):
    """Normalize an openpyxl cell value to a string, matching CSV-export behavior."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        # Render whole-number floats without a trailing ".0" (Sheets exports ints as floats sometimes).
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return str(value)


# Tab-to-file mapping with required fields and enum validation.
TAB_CONFIG = {
    "Publications": {
        "file": "data/publications.json",
        "required": ["year", "authors", "title"],
        "enums": {
            "type": [
                "indexed-jcr",
                "indexed-other",
                "non-indexed",
                "conference-derived",
                "other",
            ],
        },
    },
    "Projects": {
        "file": "data/projects.json",
        "required": ["code", "title"],
        "enums": {
            "section": ["competitive", "networks"],
            "pi_role": ["ip", "ip-national", "other-pi"],
        },
    },
    "Books": {
        "file": "data/books.json",
        "required": ["year", "authors", "title"],
        "enums": {
            "type": ["chapter", "edited-book", "proceedings"],
        },
    },
    "Conferences": {
        "file": "data/conferences.json",
        "required": ["year", "authors", "title"],
        "enums": {
            "type": ["oral", "plenary", "poster", "symposium"],
        },
    },
    "Talks": {
        "file": "data/talks.json",
        "required": ["year", "title"],
    },
    "Contracts": {
        "file": "data/contracts.json",
        "required": ["title"],
        "enums": {
            "pi_role": ["ip", "other-pi"],
        },
    },
    "Theses": {
        "file": "data/theses.json",
        "required": ["year", "author", "title"],
        "enums": {
            "type": ["phd", "master"],
            "status": ["completed", "in-progress"],
        },
    },
    "Materials": {
        "file": "data/materials.json",
        "required": ["year", "title"],
    },
}


def fetch_workbook(sheet_id):
    """Download the entire Google Sheet as an XLSX workbook."""
    url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        f"/export?format=xlsx"
    )
    try:
        resp = requests.get(url, timeout=60)
        resp.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print(f"ERROR: Failed to fetch workbook: HTTP {e.response.status_code}")
        print("  Is the Google Sheet set to 'Anyone with the link can view'?")
        sys.exit(1)
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Failed to fetch workbook: {e}")
        sys.exit(1)

    return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)


def extract_tab(workbook, tab_name):
    """Extract rows from a worksheet by name as list of dicts with lowercased headers."""
    if tab_name not in workbook.sheetnames:
        print(f"ERROR: Tab '{tab_name}' not found in workbook.")
        print(f"  Available tabs: {workbook.sheetnames}")
        sys.exit(1)

    ws = workbook[tab_name]
    rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        return []

    headers = [
        str(h).strip().lower() if h is not None else ""
        for h in header_row
    ]

    entries = []
    for row in rows:
        entry = {}
        for key, value in zip(headers, row):
            if not key:
                continue
            entry[key] = cell_to_str(value)
        if not any(entry.values()):
            continue
        entries.append(entry)
    return entries


def convert_year(entries):
    """Convert 'year' column to int where present."""
    for entry in entries:
        if "year" in entry and entry["year"]:
            try:
                entry["year"] = int(entry["year"])
            except (ValueError, TypeError):
                pass  # Leave as-is; validation will catch issues
    return entries


def sort_by_year(entries):
    """Sort entries by year descending."""
    def year_key(e):
        y = e.get("year", 0)
        return y if isinstance(y, int) else 0
    return sorted(entries, key=year_key, reverse=True)


def validate(entries, tab_name, config):
    """Validate required fields and enum values. Return list of errors."""
    errors = []
    required = config.get("required", [])
    enums = config.get("enums", {})

    for i, entry in enumerate(entries):
        row_num = i + 2  # 1-indexed, skip header row
        for field in required:
            val = entry.get(field, "")
            if val == "":
                errors.append(
                    f"{tab_name} row {row_num}: missing required field '{field}'"
                )
        for field, allowed in enums.items():
            val = entry.get(field, "")
            if val and val not in allowed:
                errors.append(
                    f"{tab_name} row {row_num}: invalid '{field}' value "
                    f"'{val}' (allowed: {allowed})"
                )
    return errors


def write_json(filepath, entries):
    """Write entries to a JSON file."""
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(entries, f, indent=2, ensure_ascii=False)
        f.write("\n")
    print(f"  Wrote {len(entries)} entries to {filepath}")


def load_workbook_from_file(path):
    """Load an XLSX workbook from a local file."""
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        metavar="PATH",
        help="Use a local XLSX file instead of downloading from Google Sheets",
    )
    args = parser.parse_args()

    if args.xlsx:
        print(f"Loading workbook from {args.xlsx}...")
        workbook = load_workbook_from_file(args.xlsx)
    else:
        sheet_id = os.environ.get("GOOGLE_SHEET_ID")
        if not sheet_id:
            print("ERROR: GOOGLE_SHEET_ID environment variable is not set.")
            sys.exit(1)
        if not re.fullmatch(r"[A-Za-z0-9_-]+", sheet_id):
            print("ERROR: GOOGLE_SHEET_ID contains invalid characters.")
            sys.exit(1)

        print("Fetching workbook...")
        workbook = fetch_workbook(sheet_id)

    print(f"  Tabs available: {workbook.sheetnames}")

    all_errors = []

    for tab_name, config in TAB_CONFIG.items():
        print(f"Processing tab: {tab_name}")
        entries = extract_tab(workbook, tab_name)
        entries = convert_year(entries)
        entries = sort_by_year(entries)

        errors = validate(entries, tab_name, config)
        if errors:
            print(f"  Fields found: {list(entries[0].keys()) if entries else '(none)'}")
            if entries:
                print(f"  First entry: {entries[0]}")
        all_errors.extend(errors)

        # Write files even if validation fails (for debugging)
        write_json(config["file"], entries)

    if all_errors:
        print(f"\nValidation failed with {len(all_errors)} error(s):")
        for err in all_errors:
            print(f"  - {err}")
        sys.exit(1)

    print("\nAll tabs processed successfully.")


if __name__ == "__main__":
    main()
